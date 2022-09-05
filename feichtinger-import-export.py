from ast import arg
from io import StringIO, TextIOWrapper
import numbers
import openpyxl
import re
import argparse
import sys
import csv

from datetime import *
today = datetime.today()
current_week_nr = today.strftime("%U")
print("Aktuelle Woche:",current_week_nr)

parser = argparse.ArgumentParser(description="Convert Bestellungen to Foodsoft")
parser.add_argument('-i', '--in_file', help="Input file vom Lieferanten")
parser.add_argument('-o', '--out_file' ,nargs='?', type=argparse.FileType('w'),default=sys.stdout, help="Das csv zum Upload in die foodsoft")
parser.add_argument('-b', '--bestellung', nargs='?', type=argparse.FileType('r',encoding='iso-8859-1'), help="Das csv mit den Bestellungen")
parser.add_argument('-w', '--week', default=current_week_nr, type=int, help="Defaults to the current week number")
parser.add_argument('-d', '--debug', default=False, type=bool, help="Debug on/off")
parser.add_argument('-f', '--filter', type=str, help="Regex Filter for artikel name, which should not be on the csv to foodsoft list")

args = parser.parse_args()
print("in_file >",type(args.in_file), file=sys.stderr)
print("out_file >",type(args.out_file), file=sys.stderr)
print("bestellung >",type(args.bestellung), file=sys.stderr)

if type(args.in_file) != str: 
    parser.print_help()
    exit

def artikel2Bestellung():
    print("Starte import")
    bestellt_filename = "KW{}_{}".format(args.week,args.in_file)
    #shutil.copyfile(args.in_file,bestellt_filename)
    mywb = openpyxl.load_workbook(args.in_file)
    sheet = mywb.active
    if args.debug: print( sheet )
    summe = 0
    dict_summe = {}
    csv_in = csv.reader(args.bestellung,delimiter=';')

    for line in csv_in:
        if args.debug: print(line)
        if not re.match('\d+',line[1]):
            continue
        name = line[2]
        fs_einheit = line[3]
        fs_preis = float(re.match('([\d,]+)+',line[5])[1].replace(',','.'))
        fs_preisgesammt = float(re.match('([\d,]+)+',line[6])[1].replace(',','.'))
        
        nr = int(line[1])
        fs_menge = int(line[0])
        
        summe = summe + ( fs_menge * fs_preis )
        
        print(nr,name,fs_preis,fs_menge)
        for x in range (11,100):
            if sheet.cell(row=x,column=1).value == nr:
                print("<----- Found Record Nr:{} in Row: {} ------>".format(nr,x))
                einheit = sheet.cell(row=x,column=3).value
                preis = sheet.cell(row=x,column=5).value
                print("Preis: {}, Einheit: {}".format(preis,einheit))
                if einheit == 'kg' and fs_einheit == '500g':
                    fs_menge = fs_menge / 2
                    
                curr_val = sheet.cell(row=x,column=8).value
                if args.debug: print( "curr_val:",type(curr_val),"Val:",curr_val)
                if type(curr_val) in [float,int] and curr_val > 0:
                    fs_menge = fs_menge + curr_val
                    print( "Wert für {} wird erhöht von {} auf {}".format(name,curr_val,fs_menge))
                              
                sheet.cell(row=x,column=8).value = fs_menge
                if args.debug: print("Schreibe Menge {}".format(fs_menge))
                
                continue;

    mywb.save(bestellt_filename)
    print( "Saved {}, Summe: {}".format(bestellt_filename,summe))


def bestell2artikel():
    
    mywb = openpyxl.load_workbook(args.in_file)

    sheet = mywb.active
    print( sheet )

    # Feichtinger xls
    # Artikel NR,	Artikel,	EH,	Zusatz,	Preis,	incl. MwSt. %,	Bemerkung
    # 11 first row
    #csv = open("feichtinger_artikel.csv", "w+")
    csv = args.out_file
    print('verf.','Bestellnummer','Name','Notiz','Produzent','Herkunft','Einheit','Nettopreis','MwSt','Pfand','Gebindegröße','""','""','Kategorie',sep=';',file=csv)
    # foodsoft import csv
    # verf.;Bestellnummer;Name;Notiz;Produzent;Herkunft;Einheit;Nettopreis;MwSt;Pfand;Gebindegröße;"";"";Kategorie																									
    for x in range (11,100):

        nr = sheet.cell(row=x,column=1).value or ''
        print("NR:",nr, 'Type:',type(nr), file=sys.stderr)
        # ignoriere alle ohne artikel nr und alle artikel nummer ab 100000 ( gemüse ist hoffentlich immer unter 100000)
        if not type(nr) == int or nr > 99999: continue

        name = sheet.cell(row=x,column=2).value
        name = re.sub('"','',name)
        einheit = sheet.cell(row=x,column=3).value
        zusatz = sheet.cell(row=x,column=4).value
        preis = sheet.cell(row=x,column=5).value
        mwst = sheet.cell(row=x,column=6).value
        comment = sheet.cell(row=x,column=7).value or ''
        # 
        comment = re.sub('"','',comment)

        if args.filter and re.match(args.filter,name):
            print("Filtered Name: ",name," with filter ",args.filter)
            continue

        if args.filter and re.match(args.filter,comment):
            print("Filtered Comment: ",comment," with filter ",args.filter)
            continue

        print(x,nr,name,einheit,zusatz,preis,mwst,comment,sep=";", file=sys.stderr)
        if re.match( '.*zukauf.*',comment,flags=re.IGNORECASE):
            print("...skipped", file=sys.stderr)
            continue

        fs_bestellnummer = nr
        fs_name = name
        fs_notiz = comment
        fs_herkunft = 'AT'
        fs_produzent = ''
        fs_herkunft = ''
        fs_pfand = '0.0'

        m = re.match( ".*((\d+) kg)+.*" , name)
        #print(m)
        if m:
            #print(m.groups())
            fs_einheit = m.group(2)+'kg'
        else:
            fs_einheit = einheit

        print('---->',fs_einheit, file=sys.stderr)

        fs_nettopreis = preis
        # mwst leer
        fs_mwst = '0.0'
        fs_gebinde = '1'
        if einheit == 'Fl.':
            fs_kategorie = 'Getränke'
        else:
            fs_kategorie = 'Gemüse'

     #   if fs_einheit == '1kg':
     #       fs_name_500g = re.sub('1 kg','500g',fs_name)
     #       if fs_name == fs_name_500g:
     #           fs_name_500g = fs_name + " 500g"
     #       print( '""',fs_bestellnummer,fs_name_500g,fs_notiz,fs_produzent,fs_herkunft,'500g',fs_nettopreis/2,fs_mwst,fs_pfand,fs_gebinde,'""','""',fs_kategorie, sep=';',file=csv)

        if fs_einheit == 'kg' and not re.match( '.*(lose|kg$)' , fs_name) :
            fs_name_500g = re.sub('kg','500g',fs_name)
    # name soll gleich bleiben, nur die einheit von 1kg auf 500g
    #        is_name == ff fs_name_500g:
    #            fs_name_500g = fs_name + " 500g"
            print( '""',fs_bestellnummer,fs_name_500g + '(500g)',fs_notiz,fs_produzent,fs_herkunft,'500g',fs_nettopreis/2,fs_mwst,fs_pfand,fs_gebinde,'""','""',fs_kategorie, sep=';',file=csv)
        else:
            print( '""',fs_bestellnummer,fs_name+ '('+fs_einheit+')',fs_notiz,fs_produzent,fs_herkunft,fs_einheit,fs_nettopreis,fs_mwst,fs_pfand,fs_gebinde,'""','""',fs_kategorie, sep=';',file=csv)


    print("fertig", file=sys.stderr)


if type(args.in_file) == str and type(args.bestellung) == TextIOWrapper:
    artikel2Bestellung()
elif type(args.in_file) == str and type(args.out_file) == TextIOWrapper:
    bestell2artikel()



